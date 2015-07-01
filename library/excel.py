from openpyxl import Workbook
from openpyxl import load_workbook

class Workbook():
  def __init__(self):
    self.wb = load_workbook('Default.xlsx')
    self.ws = self.wb.active
    self.ws.title = "Registration"

  def save_workbook(self, event):
    self.wb.save(event.name + '.xlsx') 

  def save_tournament_data(self, event):
    self.ws['A1'] = event.name
    self.ws['C1'] = "TO'd by: " + event.TO_name
    self.ws['E1'] = "Date: " + event.date
    self.ws['G1'] = "Total Entrants"
    self.ws['I1'] = event.number_of_entrants
    self.ws['A2'] = "Number"
    self.ws['B2'] = "Tag"
    self.ws['D2'] = "Name"
    self.ws['F2'] = "Location"
    self.ws['G2'] = "Total Owed"
    counter = 72
    for tournament in event.tournaments:
      self.ws[str(unichr(counter)) + '2'] = (tournament.name + ": " 
                                             "" + str(tournament.price))
      self.ws[str(unichr(counter)) + '3'] = tournament.total
      counter += 1

  def save_entrant_data(self, event, entrant):
    self.ws['A' + str(entrant.number + 3)] = entrant.number
    self.ws['B' + str(entrant.number + 3)] = entrant.tag
    self.ws['D' + str(entrant.number + 3)] = entrant.name
    self.ws['F' + str(entrant.number + 3)] = entrant.location
    self.ws['G' + str(entrant.number + 3)] = entrant.amount_owed
    counter = 72
    for tournament in event.tournaments:
      if tournament in entrant.tournaments_parents:
        self.ws[str(unichr(counter)) + str(entrant.number + 3)] = \
                                                        tournament.price
      else:
        self.ws[str(unichr(counter)) + str(entrant.number + 3)] = 0
      counter += 1 