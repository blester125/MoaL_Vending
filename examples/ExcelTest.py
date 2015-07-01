#!/usr/bin/python

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys

def saveTournamentData(ws, event):
  ws['A1'] = event.name
  ws['C1'] = "TO'd by: " + event.TO_name
  ws['E1'] = "Date: " + event.date
  ws['G1'] = "Total Entrants"
  ws['I1'] = event.number_of_entrants
  ws['A2'] = "Number"
  ws['B2'] = "Tag"
  ws['D2'] = "Name"
  ws['F2'] = "Location"
  ws['G2'] = "Total Owed"
  counter = 72
  for tournament in event.events:
    ws[str(unichr(counter))+'2'] = tournament.name + ": " + str(tournament.price)
    ws[str(unichr(counter))+'3'] = tournament.total
    counter += 1
  
def saveEntrantData(ws, event, entrant):
  ws['A' + str(entrant.number + 3)] = entrant.number
  ws['B' + str(entrant.number + 3)] = entrant.tag
  ws['D' + str(entrant.number + 3)] = entrant.name
  ws['F' + str(entrant.number + 3)] = entrant.location
  ws['G' + str(entrant.number + 3)] = entrant.amount_owed
  counter = 72
  for tournament in event.events:
    if tournament in entrant.events:
      ws[str(unichr(counter))+str(entrant.number + 3)] = tournament.price
    else:
      ws[str(unichr(counter))+str(entrant.number + 3)] = 0
    counter += 1
    
def waitForPayment(amount_owed):
  pay = raw_input("Press Enter then they Pay")
  return True   
 
#event Class info and what is at the event and who runs it
class Event():
  #TO_name
  #date
  #number_of_entrents
  def __init__(self):
    #list of tournament objects
    self.events = []
    self.entrants = []
    
class Entrant():
  #name
  #tag
  #number
  #amount_owed
  #location
  #Teammate
  def __init__(self):
    self.events = []
    
class Tournament():
  def __init__(self):
    self.name = ""
    self.price = 0
    self.total = 0
    self.team = False

def setMoaLData(event):
  event.TO_name = "Brian Lester & Kofi Osei"
  number = raw_input("What number MoaL is this? ")
  event.name = "MoaL " + str(number)
  event.date = raw_input("What is the date? ")
  venue = Tournament()
  venue.name = "Venue"
  venue.price = 1
  Singles = Tournament()
  Singles.name = "Singles"
  Singles.price = 1
  Doubles = Tournament()
  Doubles.name = "Doubles"
  Doubles.price = 1
  event.events.append(venue)
  event.events.append(Singles)
  event.events.append(Doubles)  
    
def main():    
  #open a blank workbook to write into
  wb = load_workbook('Default.xlsx')
  ws = wb.active
  ws.title = "Registration"
  event = Event()
  event.number_of_entrants = 0
  moal = raw_input("Is this for MoaL? (y/n)")
  if moal == 'y' or moal == "Y":
    setMoaLData(event)
  else:
    #Collect TO info
    event.TO_name = raw_input("Enter the name of the T.O. (You): ")
    event.name = raw_input("Enter the name of the event: ")
    event.date = raw_input("Enter the date of the tournament: ")
    venue = raw_input("Are you charging a Venue Fee? (y/n)")
    if venue == 'y' or venue == "Y":
      tournament = Tournament()
      cost = raw_input("How much is the Venue Fee? ")
      tournament.name = "Venue"
      tournament.price = int(cost)
      event.events.append(tournament)
    while True:
      tournament_name = raw_input("Enter the name of a Singles Event you are running \n(Enter nothing to finish entering Events) ")
      if tournament_name == "":
        break
      tournament = Tournament()
      tournament.name = tournament_name
      event.events.append(tournament)
      cost = raw_input("Enter the Cost of this Tournament: ")
      tournament.price = int(cost)
    while True:
      tournament_name = raw_input("Enter the name of a Doubles Event you are running \n(Enter nothing to finish entering Events) ")
      if tournament_name == "":
        break
      tournament = Tournament()
      tournament.name = tournament_name
      event.events.append(tournament)
      cost = raw_input("Enter the Cost of this Tournament: ")
      tournament.price = int(cost)
      tournament.team = True
  
  for i in range(0, 2):
    os.system("cls")
    entrant = Entrant()
    event.number_of_entrants += 1
    event.entrants.append(entrant)
    entrant.number = event.number_of_entrants
    entrant.amount_owed = 0
    entrant.tag = raw_input("Enter your gamer tag: ")
    entrant.name = raw_input("Enter your name: ")
    entrant.location = raw_input("Enter where are you from: ")
    for tournament in event.events:
      #if time > tournament.reg_time
      #  continue
      if tournament.name == "Venue":
        entrant.events.append(tournament)
        entrant.amount_owed = tournament.price
        tournament.total += tournament.price
      else:
        tourny = raw_input('Are you entering ' + tournament.name + ' ? (y/n)')
        if tourny == 'y' or tourny == 'Y':
          entrant.events.append(tournament)
          entrant.amount_owed += int(tournament.price)
          tournament.total += int(tournament.price)
    print "You owe: " + str(entrant.amount_owed)
    waitForPayment(entrant.amount_owed)
    print "You Payed!"
    saveEntrantData(ws, event, entrant)
  
  saveTournamentData(ws, event)
  wb.save(event.name + '.xlsx')
  
if __name__ == "__main__":
  main()
